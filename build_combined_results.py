"""Build Final_Combined_Results.xlsx — one workbook with both CS and Slump.

Merges CompressiveStrength_ML_Results_v3.xlsx and Slump_ML_Results_v2.xlsx
into a single workbook with clear section labels. For one-stop submission.
"""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, "/Users/fakhirhassan/Desktop/CivilFYP")

CS_PATH = "/Users/fakhirhassan/Desktop/CivilFYP/CompressiveStrength_ML_Results_v3.xlsx"
SLUMP_PATH = "/Users/fakhirhassan/Desktop/CivilFYP/Slump_ML_Results_v2.xlsx"
OUT = "/Users/fakhirhassan/Desktop/CivilFYP/Final_Combined_Results.xlsx"

CS_FILL = PatternFill("solid", fgColor="1F4E78")  # navy
SLUMP_FILL = PatternFill("solid", fgColor="C0504D")  # red
SUMMARY_FILL = PatternFill("solid", fgColor="385723")  # green
WHITE = Font(color="FFFFFF", bold=True, size=12)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell, fill, font=WHITE):
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def copy_sheet(src_ws, dst_wb, new_name, prefix_color=None):
    """Copy a sheet from one workbook to another, preserving values + simple styles."""
    from openpyxl.cell.cell import MergedCell
    new_ws = dst_wb.create_sheet(new_name[:31])
    for row in src_ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            try:
                if cell.has_style:
                    new_cell.font = Font(name=cell.font.name, size=cell.font.size,
                                         bold=cell.font.bold, italic=cell.font.italic,
                                         color=cell.font.color)
                    new_cell.fill = PatternFill(fill_type=cell.fill.fill_type,
                                                fgColor=cell.fill.fgColor)
                    new_cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        wrap_text=cell.alignment.wrap_text)
                    new_cell.border = BORDER
                    new_cell.number_format = cell.number_format
            except Exception:
                pass
    # Copy column widths
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            new_ws.column_dimensions[col_letter].width = dim.width
    # Copy row heights
    for row_idx, dim in src_ws.row_dimensions.items():
        if dim.height:
            new_ws.row_dimensions[row_idx].height = dim.height
    # Copy merged cells
    for merged_range in src_ws.merged_cells.ranges:
        try:
            new_ws.merge_cells(str(merged_range))
        except Exception:
            pass
    return new_ws


def build_summary_sheet(wb):
    """A friendly cover/summary sheet with the headline numbers from both projects."""
    import pandas as pd

    ws = wb.create_sheet("📊 Summary", 0)

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "GEOPOLYMER CONCRETE — COMBINED ML RESULTS"
    style_header(c, fill=SUMMARY_FILL, font=Font(color="FFFFFF", bold=True, size=14))
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A3:F3")
    ws["A3"] = ("This workbook combines the machine-learning results for both prediction targets: "
                "28-day compressive strength (in MPa) and slump (in mm). All models tested under "
                "5-fold cross-validation × 5 repeats, plus a stricter group-aware split.")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[3].height = 50

    # Section: CS
    ws.merge_cells("A5:F5")
    style_header(ws["A5"], fill=CS_FILL)
    ws["A5"] = "COMPRESSIVE STRENGTH (28-day, MPa)"

    headers = ["Rank", "Model", "Random-KF Test R²", "GroupKFold Test R²", "Gap (Train−Test)", "Notes"]
    for j, h in enumerate(headers, 1):
        style_header(ws.cell(row=6, column=j, value=h), fill=CS_FILL)

    cs_rows = [
        (1, "Stacked Ensemble", 0.710, 0.523, 0.250, "Best overall — combines 4 diverse base models"),
        (2, "ExtraTrees", 0.698, 0.538, 0.268, "Best single tree-ensemble"),
        (3, "CatBoost", 0.680, 0.519, 0.305, "Original report's top model"),
        (4, "Gradient Boosting", 0.679, 0.481, 0.319, "Reliable boosting"),
        (5, "LightGBM", 0.666, 0.444, 0.306, "Fast and accurate"),
    ]
    for i, row in enumerate(cs_rows, start=7):
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.alignment = Alignment(horizontal="left" if j in (2, 6) else "center", vertical="center")
            cell.border = BORDER
            if isinstance(v, float):
                cell.number_format = "0.000"

    # Section: Slump
    ws.merge_cells("A13:F13")
    style_header(ws["A13"], fill=SLUMP_FILL)
    ws["A13"] = "SLUMP (mm)"

    for j, h in enumerate(headers, 1):
        style_header(ws.cell(row=14, column=j, value=h), fill=SLUMP_FILL)

    slump_rows = [
        (1, "Gradient Boosting", 0.861, 0.481, 0.138, "Best single model on random-KF"),
        (2, "XGBoost", 0.857, 0.531, 0.141, "Close second"),
        (3, "LightGBM", 0.835, 0.662, 0.306, "Best honest GroupKFold of the boosting trio"),
        (4, "CatBoost", 0.834, 0.540, 0.272, "Solid all-rounder"),
        (5, "Stacked Ensemble", 0.816, 0.674, 0.152, "Best honest GroupKFold overall"),
    ]
    for i, row in enumerate(slump_rows, start=15):
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.alignment = Alignment(horizontal="left" if j in (2, 6) else "center", vertical="center")
            cell.border = BORDER
            if isinstance(v, float):
                cell.number_format = "0.000"

    # Methodology note
    ws.merge_cells("A21:F21")
    ws["A21"] = "METHODOLOGY"
    style_header(ws["A21"], fill=SUMMARY_FILL)
    ws.row_dimensions[21].height = 22

    notes = [
        "Random-KF Test R² = mean test R² across 5-fold cross-validation, 5 repeats (25 averaged folds). Comparable to typical published results.",
        "GroupKFold Test R² = honest test where the same base mix recipe NEVER appears in both training and test set. Harder bar, more rigorous.",
        "Gap (Train−Test) = how much accuracy drops from training data to held-out test data. Smaller is better — large gap means the model just memorized the training set.",
        "All models tuned via Optuna Bayesian optimization (100 trials per boosting model, 50 per smaller model), optimising for GroupKFold validation R².",
        "Compressive Strength dataset: 353 rows × 15 features. Slump dataset: 84 rows × 15 features.",
        "Stacked Ensemble = SVR + ExtraTrees + XGBoost + HistGradientBoosting (CS) or HGB + ExtraTrees + GP + KNN (Slump), with Ridge meta-learner.",
    ]
    for i, n in enumerate(notes, start=22):
        ws.merge_cells(start_row=i, end_row=i, start_column=1, end_column=6)
        cell = ws.cell(row=i, column=1, value=n)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 36

    # Improvement section
    base_row = 22 + len(notes) + 1
    ws.merge_cells(start_row=base_row, end_row=base_row, start_column=1, end_column=6)
    ws.cell(row=base_row, column=1, value="IMPROVEMENT vs ORIGINAL REPORTS")
    style_header(ws.cell(row=base_row, column=1), fill=SUMMARY_FILL)
    ws.row_dimensions[base_row].height = 22

    headers2 = ["Target", "Original Test R²", "Our Test R²", "Δ", "Original Gap", "Our Gap"]
    for j, h in enumerate(headers2, 1):
        style_header(ws.cell(row=base_row+1, column=j, value=h), fill=SUMMARY_FILL)

    impr = [
        ("Compressive Strength", 0.6749, 0.7102, "+5.2%", 0.32, 0.25),
        ("Slump", 0.97, 0.861, "−11% (more honest)", 0.02, 0.138),
    ]
    for i, row in enumerate(impr, start=base_row+2):
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.alignment = Alignment(horizontal="left" if j == 1 else "center")
            cell.border = BORDER
            if isinstance(v, float):
                cell.number_format = "0.000"

    # Set column widths
    widths = {1: 8, 2: 30, 3: 18, 4: 18, 5: 18, 6: 60}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    return ws


def main():
    print("Loading source workbooks …")
    cs_wb = load_workbook(CS_PATH)
    sl_wb = load_workbook(SLUMP_PATH)

    out = Workbook()
    out.remove(out.active)

    # Cover/summary sheet first
    print("Building summary sheet …")
    build_summary_sheet(out)

    # CS section
    print("Copying compressive-strength sheets …")
    for sheet_name in cs_wb.sheetnames:
        new_name = f"CS_{sheet_name}"
        copy_sheet(cs_wb[sheet_name], out, new_name)
        print(f"  CS_{sheet_name}")

    # Slump section
    print("Copying slump sheets …")
    for sheet_name in sl_wb.sheetnames:
        new_name = f"SL_{sheet_name}"
        copy_sheet(sl_wb[sheet_name], out, new_name)
        print(f"  SL_{sheet_name}")

    out.save(OUT)
    print(f"\nSaved -> {OUT}")
    print(f"Total sheets: {len(out.sheetnames)}")


if __name__ == "__main__":
    main()
