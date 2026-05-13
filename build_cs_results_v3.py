"""Build CompressiveStrength_ML_Results_v3.xlsx using the FULL 353-row dataset.

Same layout as v2 but evaluates against tuned params from cs_full_phase_results.xlsx.
"""
from __future__ import annotations
import sys, math
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, "/Users/fakhirhassan/Desktop/CivilFYP")
import civil_utils as cu  # noqa
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.model_selection import GroupKFold, KFold

OUT = "/Users/fakhirhassan/Desktop/CivilFYP/CompressiveStrength_ML_Results_v3.xlsx"
PHASE_LOG = "/Users/fakhirhassan/Desktop/CivilFYP/cs_full_phase_results.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
LIGHT_FILL = PatternFill("solid", fgColor="DEEBF7")
METRIC_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell, fill=HEADER_FILL, font=WHITE):
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def style_cell(cell, bold=False, fill=None, align="center"):
    if bold:
        cell.font = BOLD
    if fill is not None:
        cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = BORDER


def autosize(ws, min_width=10, max_width=22):
    for col_idx, col in enumerate(ws.columns, 1):
        length = min_width
        for cell in col:
            v = cell.value
            if v is None: continue
            length = max(length, len(str(v)) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(length, max_width)


def all_metrics(y_true, y_pred):
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)
    err = y_pred - y_true
    mae = float(mean_absolute_error(y_true, y_pred))
    mse = float(mean_squared_error(y_true, y_pred))
    rmse = float(math.sqrt(mse))
    r2 = float(r2_score(y_true, y_pred))
    denom_lmi = np.sum(np.abs(y_true - y_true.mean()))
    lmi = float(1.0 - np.sum(np.abs(err)) / denom_lmi) if denom_lmi else float("nan")
    eae = float(np.mean(np.abs(err) / np.maximum(np.abs(y_true), 1e-9)))
    vaf = float(100 * (1 - np.var(err) / np.var(y_true))) if np.var(y_true) else float("nan")
    std = float(np.std(err, ddof=0))
    return {"MAE":mae,"MSE":mse,"RMSE":rmse,"R2":r2,"LMI":lmi,"EAE":eae,"VAF":vaf,"STD":std}


METRIC_ORDER = ["MAE","MSE","RMSE","R2","LMI","EAE","VAF","STD"]


def setup_data():
    df = cu.cs_full_dataset()
    Xeng = cu.engineer_features(df, mode="engineered", target=cu.CS_TARGET)
    X12 = Xeng.drop(columns=[cu.CS_TARGET])
    y = Xeng[cu.CS_TARGET].values
    # Use the actual perm-imp ranking from the new run for engineered_9
    imp = pd.read_excel(PHASE_LOG, sheet_name="Phase2_perm_importance")
    drop9 = imp.tail(3)["feature"].tolist()
    X9 = X12.drop(columns=[c for c in drop9 if c in X12.columns])
    Xraw = cu.engineer_features(df, mode="raw", target=cu.CS_TARGET).drop(columns=[cu.CS_TARGET])
    groups = cu.assign_group_ids(Xraw)
    return df, X12, X9, Xraw, y, groups, drop9


def best_params_lookup():
    out = {}
    for sheet in ("Phase3_best_params", "Phase4_best_params"):
        src = pd.read_excel(PHASE_LOG, sheet_name=sheet)
        for _, r in src.iterrows():
            fset = r["feature_set"]; model = r["model"]
            params = {}
            for c in src.columns:
                if c in ("feature_set","model"): continue
                v = r[c]
                if isinstance(v, float) and np.isnan(v): continue
                if isinstance(v, float) and float(v).is_integer() and c in (
                    "n_estimators","max_depth","min_samples_split","min_samples_leaf",
                    "iterations","depth","min_data_in_leaf","num_leaves","min_child_weight",
                    "max_iter","max_leaf_nodes","n_neighbors","p","n_components","degree",
                ):
                    v = int(v)
                params[c] = v
            out[(fset, model)] = params
    return out


# Same model list as v2 — keeps comparability
MODEL_LIST = [
    ("Linear Regression", None,                "raw"),
    ("Random Forest",      "Random Forest",      None),
    ("Gradient Boosting",  "Gradient Boosting",  None),
    ("AdaBoost",           "AdaBoost",           None),
    ("CatBoost",           "CatBoost",           None),
    ("LightGBM",           "LightGBM",           None),
    ("XGBoost",            "XGBoost",            None),
    ("SVR (RBF)",          "SVR_RBF",            None),
    ("ExtraTrees",         "ExtraTrees",         None),
    ("HistGradientBoosting","HistGradientBoosting", None),
    ("Gaussian Process",   "GaussianProcess",    None),
    ("Stacked Ensemble",   "__STACK__",          "engineered_12"),
]


def make_factory(model_internal, params_lookup, fset_name):
    if model_internal == "Linear Regression" or model_internal is None:
        from sklearn.linear_model import LinearRegression
        return lambda: LinearRegression()
    if model_internal == "__STACK__":
        from sklearn.linear_model import Ridge
        from sklearn.pipeline import Pipeline
        from sklearn.preprocessing import StandardScaler
        bases = {}
        for nm in ("SVR_RBF","ExtraTrees","XGBoost","HistGradientBoosting"):
            params = params_lookup.get((fset_name, nm), {})
            bases[nm] = cu.make_tuned_factory(nm, params)
        meta = lambda: Pipeline([("sc", StandardScaler()),("m", Ridge(alpha=1.0, random_state=42))])
        return ("STACK", bases, meta)
    params = params_lookup.get((fset_name, model_internal), {})
    return cu.make_tuned_factory(model_internal, params)


def representative_fold_preds(factory, X, y, groups=None, regime="random"):
    if regime == "random":
        kf = KFold(n_splits=5, shuffle=True, random_state=0)
        splits = list(kf.split(X, y))
    else:
        splits = list(GroupKFold(n_splits=5).split(X, y, groups))
    rng = np.random.default_rng(0)
    fold_results = []
    for trainval_idx, test_idx in splits:
        if regime == "random":
            shuffled = rng.permutation(trainval_idx)
            n_val = max(1, int(round(len(shuffled) * 0.2)))
            val_idx = shuffled[:n_val]; train_idx = shuffled[n_val:]
        else:
            inner_groups = groups[trainval_idx]
            unique = np.unique(inner_groups)
            rng.shuffle(unique)
            n_val_groups = max(1, int(round(len(unique) * 0.2)))
            val_groups = set(unique[:n_val_groups].tolist())
            mask = np.isin(inner_groups, list(val_groups))
            val_idx = trainval_idx[mask]; train_idx = trainval_idx[~mask]
        if len(val_idx) < 1 or len(train_idx) < 5: continue
        if isinstance(factory, tuple) and factory[0] == "STACK":
            preds = _stack_predict(factory, X, y, train_idx, val_idx, test_idx, groups, regime)
        else:
            m = factory()
            m.fit(X.iloc[train_idx].values, y[train_idx])
            preds = {
                "train": (y[train_idx], m.predict(X.iloc[train_idx].values)),
                "val":   (y[val_idx], m.predict(X.iloc[val_idx].values)),
                "test":  (y[test_idx], m.predict(X.iloc[test_idx].values)),
            }
        fold_results.append((r2_score(*preds["test"]), preds))
    if not fold_results: raise RuntimeError("no folds")
    mean_r2 = float(np.mean([fr[0] for fr in fold_results]))
    fold_results.sort(key=lambda t: abs(t[0] - mean_r2))
    return fold_results[0][1]


def _stack_predict(stack_spec, X, y, train_idx, val_idx, test_idx, groups, regime):
    _, base_factories, meta_factory = stack_spec
    n_tr = len(train_idx)
    base_names = list(base_factories.keys())
    if regime == "group" and groups is not None:
        tr_groups = groups[train_idx]
        n_inner = max(2, min(5, len(np.unique(tr_groups))))
        inner = list(GroupKFold(n_splits=n_inner).split(X.iloc[train_idx], y[train_idx], tr_groups))
    else:
        inner = list(KFold(5, shuffle=True, random_state=0).split(X.iloc[train_idx]))
    oof = np.zeros((n_tr, len(base_names)))
    val_p = np.zeros((len(val_idx), len(base_names)))
    test_p = np.zeros((len(test_idx), len(base_names)))
    train_p = np.zeros((n_tr, len(base_names)))
    Xtr_arr = X.iloc[train_idx].values; Xva_arr = X.iloc[val_idx].values; Xte_arr = X.iloc[test_idx].values
    ytr = y[train_idx]
    for j, name in enumerate(base_names):
        for itr, ite in inner:
            m = base_factories[name]()
            m.fit(Xtr_arr[itr], ytr[itr])
            oof[ite, j] = m.predict(Xtr_arr[ite])
        m_full = base_factories[name]()
        m_full.fit(Xtr_arr, ytr)
        train_p[:, j] = m_full.predict(Xtr_arr)
        val_p[:, j] = m_full.predict(Xva_arr)
        test_p[:, j] = m_full.predict(Xte_arr)
    meta = meta_factory(); meta.fit(oof, ytr)
    return {"train":(ytr, meta.predict(train_p)),
            "val":(y[val_idx], meta.predict(val_p)),
            "test":(y[test_idx], meta.predict(test_p))}


def cv_full_metrics(factory, X, y, groups=None, regime="random", n_splits=5, n_repeats=5):
    rng = np.random.default_rng(0)
    splits_list = []
    repeats = 1 if regime == "group" else n_repeats
    for repeat in range(repeats):
        if regime == "random":
            kf = KFold(n_splits=n_splits, shuffle=True, random_state=repeat)
            splits = list(kf.split(X, y))
        else:
            splits = list(GroupKFold(n_splits=n_splits).split(X, y, groups))
        for trainval_idx, test_idx in splits:
            if regime == "random":
                shuffled = rng.permutation(trainval_idx)
                n_val = max(1, int(round(len(shuffled) * 0.2)))
                val_idx = shuffled[:n_val]; train_idx = shuffled[n_val:]
            else:
                inner_groups = groups[trainval_idx]
                unique = np.unique(inner_groups)
                rng.shuffle(unique)
                n_val_groups = max(1, int(round(len(unique) * 0.2)))
                val_groups = set(unique[:n_val_groups].tolist())
                mask = np.isin(inner_groups, list(val_groups))
                val_idx = trainval_idx[mask]; train_idx = trainval_idx[~mask]
            if isinstance(factory, tuple) and factory[0] == "STACK":
                preds = _stack_predict(factory, X, y, train_idx, val_idx, test_idx, groups, regime)
            else:
                m = factory()
                m.fit(X.iloc[train_idx].values, y[train_idx])
                preds = {
                    "train": (y[train_idx], m.predict(X.iloc[train_idx].values)),
                    "val":   (y[val_idx], m.predict(X.iloc[val_idx].values)),
                    "test":  (y[test_idx], m.predict(X.iloc[test_idx].values)),
                }
            splits_list.append(preds)
    out = {}
    for split_name in ("train","val","test"):
        rows = [all_metrics(*sp[split_name]) for sp in splits_list]
        for met in METRIC_ORDER:
            out[(split_name, met)] = float(np.mean([r[met] for r in rows]))
    return out


def write_dataset_overview(wb, df, X12, y, groups):
    ws = wb.create_sheet("Dataset Overview")
    ws.merge_cells("A1:E1")
    c = ws["A1"]; c.value = "DATASET OVERVIEW — COMPRESSIVE STRENGTH (28-Day) DATA"
    style_header(c); ws.row_dimensions[1].height = 24
    headers = ["Column","Min","Max","Mean","Std Dev"]
    for j, h in enumerate(headers, 1):
        style_header(ws.cell(row=3, column=j, value=h), fill=SUB_HEADER_FILL)
    cols = list(df.columns); data = df
    for i, col in enumerate(cols, start=4):
        ws.cell(row=i, column=1, value=col)
        s = pd.to_numeric(data[col], errors="coerce")
        ws.cell(row=i, column=2, value=round(float(s.min()),4))
        ws.cell(row=i, column=3, value=round(float(s.max()),4))
        ws.cell(row=i, column=4, value=round(float(s.mean()),4))
        ws.cell(row=i, column=5, value=round(float(s.std()),4))
        for j in range(1, 6):
            style_cell(ws.cell(row=i, column=j), fill=LIGHT_FILL if i % 2 == 0 else None,
                       align="left" if j == 1 else "center")
    base = 5 + len(cols)
    items = [
        ("Source file", "compressive strength 28 days done complete .xlsx"),
        ("Total rows", int(len(df))),
        ("Rows with valid target", int(len(y))),
        ("Unique mix-design groups (strict)", int(len(np.unique(groups)))),
        ("Features (engineered_12)", int(X12.shape[1])),
        ("CV regimes reported", "Random KFold 5×5  +  GroupKFold 5×1 (strict)"),
    ]
    for i, (k, v) in enumerate(items):
        ws.cell(row=base+i, column=1, value=k).font = BOLD
        ws.cell(row=base+i, column=2, value=v)
    autosize(ws)


def write_model_performance(wb, all_metrics_random, all_metrics_group):
    ws = wb.create_sheet("Model Performance")
    title = ws.cell(row=1, column=1, value="MODEL PERFORMANCE METRICS SUMMARY — COMPRESSIVE STRENGTH (CV-averaged, full 353-row dataset)")
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=2 + len(METRIC_ORDER) * 3)
    style_header(title); ws.row_dimensions[1].height = 22

    def write_block(start_row, regime_label, metrics_by_model):
        ws.cell(row=start_row, column=1, value=regime_label).font = BOLD
        ws.cell(row=start_row, column=1).fill = SUB_HEADER_FILL
        ws.cell(row=start_row, column=1).font = WHITE
        ws.merge_cells(start_row=start_row, end_row=start_row, start_column=1, end_column=2 + len(METRIC_ORDER) * 3)
        style_cell(ws.cell(row=start_row, column=1), fill=SUB_HEADER_FILL)
        r2 = start_row + 1
        ws.cell(row=r2, column=1, value="Model")
        style_header(ws.cell(row=r2, column=1), fill=SUB_HEADER_FILL)
        for k, split in enumerate(["Train","Val","Test"]):
            c0 = 2 + k * len(METRIC_ORDER)
            ws.merge_cells(start_row=r2, end_row=r2, start_column=c0, end_column=c0 + len(METRIC_ORDER) - 1)
            style_header(ws.cell(row=r2, column=c0, value=split), fill=SUB_HEADER_FILL)
        r3 = start_row + 2
        for k in range(3):
            for j, met in enumerate(METRIC_ORDER):
                style_header(ws.cell(row=r3, column=2 + k*len(METRIC_ORDER) + j, value=met), fill=METRIC_FILL, font=BOLD)
        r = r3 + 1
        for display_name, _, _ in MODEL_LIST:
            mets = metrics_by_model.get(display_name)
            if mets is None: continue
            ws.cell(row=r, column=1, value=display_name).font = BOLD
            for k, split in enumerate(("train","val","test")):
                for j, met in enumerate(METRIC_ORDER):
                    val = mets.get((split, met))
                    cell = ws.cell(row=r, column=2 + k*len(METRIC_ORDER) + j,
                                   value=round(float(val),4) if val is not None else None)
                    cell.number_format = "0.0000"
                    style_cell(cell, fill=LIGHT_FILL if r % 2 else None)
            style_cell(ws.cell(row=r, column=1), bold=True, align="left")
            r += 1
        return r + 1

    next_row = write_block(3, "REGIME 1 — Random KFold (5 folds × 5 repeats, industry-comparable)", all_metrics_random)
    write_block(next_row, "REGIME 2 — GroupKFold strict (no mix-design leakage, honest)", all_metrics_group)
    autosize(ws, min_width=11, max_width=14)


def write_model_sheet(wb, display_name, preds_random, preds_group):
    ws = wb.create_sheet(display_name[:31])
    title = ws.cell(row=1, column=1, value=f"{display_name} — Predictions vs Actual (representative fold per regime)")
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=11)
    style_header(title)

    def write_block(start_row, label, preds):
        ws.cell(row=start_row, column=1, value=label).font = BOLD
        ws.cell(row=start_row, column=1).fill = SUB_HEADER_FILL
        ws.cell(row=start_row, column=1).font = WHITE
        ws.merge_cells(start_row=start_row, end_row=start_row, start_column=1, end_column=11)
        style_cell(ws.cell(row=start_row, column=1), fill=SUB_HEADER_FILL)
        r2 = start_row + 1
        for k, split in enumerate(("Train","Val","Test")):
            c0 = 1 + k * 4
            ws.merge_cells(start_row=r2, end_row=r2, start_column=c0, end_column=c0 + 2)
            style_header(ws.cell(row=r2, column=c0, value=split), fill=SUB_HEADER_FILL)
        r3 = start_row + 2
        for k in range(3):
            for j, h in enumerate(("Actual","Predicted","Error")):
                style_header(ws.cell(row=r3, column=1 + k*4 + j, value=h), fill=METRIC_FILL, font=BOLD)
        max_len = max(len(preds["train"][0]), len(preds["val"][0]), len(preds["test"][0]))
        for i in range(max_len):
            r = r3 + 1 + i
            for k, split in enumerate(("train","val","test")):
                actual, predicted = preds[split]
                if i < len(actual):
                    err = float(predicted[i]) - float(actual[i])
                    ws.cell(row=r, column=1 + k*4 + 0, value=round(float(actual[i]),3))
                    ws.cell(row=r, column=1 + k*4 + 1, value=round(float(predicted[i]),3))
                    ws.cell(row=r, column=1 + k*4 + 2, value=round(err,3))
                    for j in range(3):
                        style_cell(ws.cell(row=r, column=1 + k*4 + j), fill=LIGHT_FILL if i % 2 == 0 else None)
        rf = r3 + 1 + max_len + 1
        ws.cell(row=rf, column=1, value="Metric").font = BOLD
        for k, split in enumerate(("Train","Val","Test")):
            style_header(ws.cell(row=rf, column=2 + k, value=split), fill=METRIC_FILL, font=BOLD)
        for j, met in enumerate(METRIC_ORDER):
            ws.cell(row=rf + 1 + j, column=1, value=met).font = BOLD
            for k, split in enumerate(("train","val","test")):
                v = all_metrics(*preds[split])[met]
                cell = ws.cell(row=rf + 1 + j, column=2 + k, value=round(float(v),4))
                cell.number_format = "0.0000"
                style_cell(cell, fill=LIGHT_FILL if j % 2 == 0 else None)
        return rf + len(METRIC_ORDER) + 3

    next_row = write_block(3, "Random KFold (industry-comparable)", preds_random)
    write_block(next_row, "GroupKFold strict (honest)", preds_group)
    autosize(ws, min_width=12, max_width=14)


def write_methodology(wb, n_groups, drop9):
    ws = wb.create_sheet("Methodology")
    notes = [
        ("Source", "compressive strength 28 days done complete .xlsx — 353 rows, 16 columns, all targets valid."),
        ("CV regimes", "Random KFold 5×5 — industry-comparable. GroupKFold strict — honest, prevents same mix-design from sitting in both train and test."),
        ("Group definition", f"Strict key: flyash, ggbfs, na2sio3, naoh, naoh_molarity, activator_binder_ratio (rounded to 3 decimals). With this dataset → {n_groups} unique groups out of 353 rows."),
        ("Feature engineering", f"engineered_12 features: total_binder, ggbfs_fraction, total_activator, na2sio3_naoh_ratio, naoh_molarity, activator_binder_ratio, w_b_ratio, rca_fraction, r_sand_fraction, curing_temp_c, curing_time_hr, sp. engineered_9 drops {drop9} (lowest permutation importance)."),
        ("Hyperparameters", "Optuna TPE, 100 trials per boosting model + 50 trials per small-data model, optimising mean GroupKFold val R². Fresh tuning on 353 rows (not the 64-row params from v2)."),
        ("Stacked ensemble", "Bases: SVR_RBF, ExtraTrees, XGBoost, HistGradientBoosting. Meta: Ridge(alpha=1) on standardised out-of-fold base predictions."),
        ("Compared to v2", "v2 was based on 64-row subset (subset of this dataset). v3 is the full 353 rows the original report claimed."),
    ]
    ws.merge_cells("A1:B1")
    style_header(ws["A1"]); ws["A1"] = "METHODOLOGY"
    for i, (k, v) in enumerate(notes, start=3):
        ws.cell(row=i, column=1, value=k).font = BOLD
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 60
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110


def main():
    print("Loading 353-row dataset and tuned params …")
    df, X12, X9, Xraw, y, groups, drop9 = setup_data()
    params = best_params_lookup()
    feature_set_X = {"engineered_12":X12, "engineered_9":X9, "raw":Xraw}

    # Determine best feature set per model from Phase 4 leaderboard
    lb = pd.read_excel(PHASE_LOG, sheet_name="Phase4_leaderboard")
    fset_for_model = dict(zip(lb["model"], lb["feature_set"]))
    # ExtraTrees/HGB/GP/SVR may map to 'SVR_RBF' name etc. Map our internal names.
    NAME_MAP = {"SVR (RBF)":"SVR_RBF","Gaussian Process":"GaussianProcess","HistGradientBoosting":"HistGradientBoosting","ExtraTrees":"ExtraTrees"}

    metrics_random, metrics_group, sheet_preds = {}, {}, []
    for display_name, internal, fset_override in MODEL_LIST:
        print(f"  >> {display_name}")
        # decide fset
        if fset_override is not None:
            fset = fset_override
        elif internal in (None, "Linear Regression"):
            fset = "raw"
        else:
            internal_for_lookup = NAME_MAP.get(display_name, internal)
            fset = fset_for_model.get(internal_for_lookup, "engineered_12")
        X_used = feature_set_X[fset]

        if internal == "Linear Regression" or internal is None:
            factory = make_factory("Linear Regression", params, fset)
        elif display_name == "Stacked Ensemble":
            factory = make_factory("__STACK__", params, fset)
        else:
            factory = make_factory(NAME_MAP.get(display_name, internal), params, fset)

        m_rand = cv_full_metrics(factory, X_used, y, groups=None, regime="random", n_splits=5, n_repeats=5)
        m_grp = cv_full_metrics(factory, X_used, y, groups=groups, regime="group", n_splits=5, n_repeats=1)
        metrics_random[display_name] = m_rand
        metrics_group[display_name] = m_grp
        preds_rand = representative_fold_preds(factory, X_used, y, groups=None, regime="random")
        preds_grp = representative_fold_preds(factory, X_used, y, groups=groups, regime="group")
        sheet_preds.append((display_name, preds_rand, preds_grp))

    print("Writing workbook …")
    wb = Workbook(); wb.remove(wb.active)
    write_dataset_overview(wb, df, X12, y, groups)
    write_model_performance(wb, metrics_random, metrics_group)
    for display_name, pr, pg in sheet_preds:
        write_model_sheet(wb, display_name, pr, pg)
    write_methodology(wb, len(np.unique(groups)), drop9)
    wb.save(OUT)
    print(f"Saved -> {OUT}")


if __name__ == "__main__":
    main()
