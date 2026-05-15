"""Compute LOOCV inner train/val/test metrics for BoostStack and
GB_seed_ensemble, then patch the two corresponding rows in the LOOCV regime
of Slump_ML_Results_v2.xlsx.

This avoids re-running the full 90-min build for what's just two missing
val cells per slow stack. We re-use everything else in the existing xlsx.
"""
import os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from sklearn.metrics import (
    r2_score, mean_absolute_error, mean_squared_error,
)
from sklearn.model_selection import LeaveOneOut, KFold
from sklearn.linear_model import Ridge
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import GradientBoostingRegressor

import civil_utils as cu

# Reuse helpers from the builder
import build_slump_results as bsr

ROOT = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(ROOT, "Slump_ML_Results_v2.xlsx")
PHASE_LOG = os.path.join(ROOT, "slump_full_phase_results.xlsx")

METRIC_ORDER = ["MAE", "MSE", "RMSE", "R2", "LMI", "EAE", "VAF", "STD"]
LIGHT_FILL = PatternFill("solid", fgColor="DEEBF7")
BOLD = Font(bold=True)


def setup():
    df, X12, X9, Xraw, y, groups, drop9 = bsr.setup_data()
    params = bsr.best_params_lookup()
    return X9, y, groups, params


def make_boost_stack_factory(params):
    bases = {}
    for nm in ("Gradient Boosting", "LightGBM", "CatBoost", "XGBoost", "ExtraTrees"):
        ps = params.get(("engineered_9", nm), {})
        bases[nm] = cu.make_tuned_factory(nm, ps)
    meta_factory = lambda: Pipeline([
        ("sc", StandardScaler()),
        ("m", Ridge(alpha=1.0, random_state=42)),
    ])
    return ("STACK", bases, meta_factory)


def make_gb_seed_factory(params):
    gb_params = params.get(("engineered_9", "Gradient Boosting"), {})
    class _GBSeedEnsemble:
        def __init__(self):
            self.params = dict(gb_params); self.n_seeds = 10; self.models_ = []
        def fit(self, X, y):
            self.models_ = []
            for s in range(self.n_seeds):
                m = GradientBoostingRegressor(random_state=s, **self.params)
                m.fit(X, y); self.models_.append(m)
            return self
        def predict(self, X):
            return np.stack([m.predict(X) for m in self.models_], axis=0).mean(axis=0)
    return lambda: _GBSeedEnsemble()


def loocv_full_metrics(factory, X, y):
    """Same logic as build_slump_results.cv_full_metrics, regime='loocv'."""
    loo = LeaveOneOut()
    rng_loo = np.random.default_rng(0)
    oof_pred = np.zeros(len(y))
    train_actuals, train_preds, val_actuals, val_preds = [], [], [], []
    n_folds = 0
    for train_idx, test_idx in loo.split(X):
        # Inner train/val split
        shuffled = rng_loo.permutation(train_idx)
        n_iv = max(1, int(round(len(shuffled) * 0.2)))
        inner_val_idx = shuffled[:n_iv]; inner_train_idx = shuffled[n_iv:]

        if isinstance(factory, tuple) and factory[0] == "STACK":
            preds_full = bsr._stack_predict(factory, X, y, train_idx,
                                            np.array([], dtype=int), test_idx, None, "random")
            oof_pred[test_idx] = preds_full["test"][1]
            if test_idx[0] % 8 == 0:
                preds_inner = bsr._stack_predict(factory, X, y, inner_train_idx,
                                                 np.array([], dtype=int), inner_val_idx,
                                                 None, "random")
                train_actuals.extend(y[inner_train_idx])
                train_preds.extend(preds_inner["train"][1])
                val_actuals.extend(y[inner_val_idx])
                val_preds.extend(preds_inner["test"][1])
        else:
            m_full = factory()
            m_full.fit(X.iloc[train_idx].values, y[train_idx])
            oof_pred[test_idx] = m_full.predict(X.iloc[test_idx].values)
            if test_idx[0] % 8 == 0:
                m_inner = factory()
                m_inner.fit(X.iloc[inner_train_idx].values, y[inner_train_idx])
                train_actuals.extend(y[inner_train_idx])
                train_preds.extend(m_inner.predict(X.iloc[inner_train_idx].values))
                val_actuals.extend(y[inner_val_idx])
                val_preds.extend(m_inner.predict(X.iloc[inner_val_idx].values))
        n_folds += 1
        if n_folds % 10 == 0:
            print(f"      ... {n_folds}/{len(y)} folds done")

    out = {}
    test_m = bsr.all_metrics(y, oof_pred)
    train_m = bsr.all_metrics(np.array(train_actuals), np.array(train_preds))
    val_m = bsr.all_metrics(np.array(val_actuals), np.array(val_preds))
    for met in METRIC_ORDER:
        out[("train", met)] = float(train_m[met])
        out[("val", met)] = float(val_m[met])
        out[("test", met)] = float(test_m[met])
    return out


def patch_xlsx_row(wb, sheet_name, display_name, metrics):
    """Find the row whose first cell is display_name in the LOOCV regime block
    of `Model Performance`, and write the metrics into cols 2..25."""
    ws = wb[sheet_name]
    # Walk rows top-to-bottom; find regime 3 header, then the target model name beneath it
    in_loocv = False
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and "REGIME 3" in v and "LOOCV" in v:
            in_loocv = True
            continue
        if not in_loocv:
            continue
        if v == display_name:
            # Found the row. Cols: 2-9 = Train, 10-17 = Val, 18-25 = Test
            for k, split in enumerate(("train", "val", "test")):
                for j, met in enumerate(METRIC_ORDER):
                    val = metrics.get((split, met))
                    col = 2 + k * len(METRIC_ORDER) + j
                    cell = ws.cell(row=r, column=col,
                                   value=round(float(val), 4) if val is not None else None)
                    cell.number_format = "0.0000"
                    cell.alignment = Alignment(horizontal="center")
                    if r % 2:
                        cell.fill = LIGHT_FILL
            # Bold first column
            ws.cell(row=r, column=1).font = BOLD
            return True
    return False


def main():
    print("Setting up data + tuned-params lookup …")
    X9, y, groups, params = setup()
    print(f"  X9 shape: {X9.shape}, y len {len(y)}")

    targets = [
        ("BoostStack", make_boost_stack_factory(params)),
        ("GB_seed_ensemble", make_gb_seed_factory(params)),
    ]

    new_metrics = {}
    for name, factory in targets:
        print(f"\n>> LOOCV inner train/val for {name} …")
        m = loocv_full_metrics(factory, X9, y)
        new_metrics[name] = m
        print(f"   {name}: train_R2={m[('train','R2')]:.4f}  val_R2={m[('val','R2')]:.4f}  test_R2={m[('test','R2')]:.4f}")

    print("\nPatching xlsx …")
    wb = openpyxl.load_workbook(XLSX)
    for name, m in new_metrics.items():
        ok = patch_xlsx_row(wb, "Model Performance", name, m)
        print(f"  {name}: {'patched' if ok else 'NOT FOUND'}")
    wb.save(XLSX)
    print(f"Saved -> {XLSX}")


if __name__ == "__main__":
    main()
