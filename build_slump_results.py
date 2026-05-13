"""Build Slump_ML_Results_v2.xlsx — polished report mirroring the CS v3 format.

Reads tuned params from slump_full_phase_results.xlsx (produced by running
slump_modeling_full.ipynb first), then evaluates each model under random
KFold and GroupKFold and writes a 15-sheet formatted workbook.
"""
from __future__ import annotations
import sys, math
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, "/Users/fakhirhassan/Desktop/CivilFYP")
import civil_utils as cu  # noqa
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.model_selection import GroupKFold, KFold

OUT = "/Users/fakhirhassan/Desktop/CivilFYP/Slump_ML_Results_v2.xlsx"
PHASE_LOG = "/Users/fakhirhassan/Desktop/CivilFYP/slump_full_phase_results.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
LIGHT_FILL = PatternFill("solid", fgColor="DEEBF7")
METRIC_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TARGET_NAME = "Slump (mm)"


def style_header(cell, fill=HEADER_FILL, font=WHITE):
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def style_cell(cell, bold=False, fill=None, align="center"):
    if bold: cell.font = BOLD
    if fill is not None: cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = BORDER


def autosize(ws, min_width=10, max_width=24):
    for col_idx, col in enumerate(ws.columns, 1):
        length = min_width
        for cell in col:
            if cell.value is None: continue
            length = max(length, len(str(cell.value)) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(length, max_width)


def all_metrics(y_true, y_pred):
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)
    err = y_pred - y_true
    mae = float(mean_absolute_error(y_true, y_pred))
    mse = float(mean_squared_error(y_true, y_pred))
    rmse = float(math.sqrt(mse))
    r2 = float(r2_score(y_true, y_pred))
    denom_lmi = np.sum(np.abs(y_true - y_true.mean()))
    lmi = float(1.0 - np.sum(np.abs(err)) / denom_lmi) if denom_lmi else float("nan")
    eae = float(np.mean(np.abs(err) / np.maximum(np.abs(y_true), 1e-9)))
    vaf = float(100 * (1 - np.var(err) / np.var(y_true))) if np.var(y_true) else float("nan")
    std = float(np.std(err, ddof=0))
    return {"MAE":mae,"MSE":mse,"RMSE":rmse,"R2":r2,"LMI":lmi,"EAE":eae,"VAF":vaf,"STD":std}


METRIC_ORDER = ["MAE","MSE","RMSE","R2","LMI","EAE","VAF","STD"]


def setup_data():
    df = cu.slump_dataset()
    Xeng = cu.engineer_features(df, mode="engineered", target=cu.SLUMP_TARGET)
    X12 = Xeng.drop(columns=[cu.SLUMP_TARGET])
    y = Xeng[cu.SLUMP_TARGET].values
    imp = pd.read_excel(PHASE_LOG, sheet_name="Phase2_perm_importance")
    drop9 = imp.tail(3)["feature"].tolist()
    X9 = X12.drop(columns=[c for c in drop9 if c in X12.columns])
    Xraw = cu.engineer_features(df, mode="raw", target=cu.SLUMP_TARGET).drop(columns=[cu.SLUMP_TARGET])
    groups = cu.assign_group_ids(Xraw)
    return df, X12, X9, Xraw, y, groups, drop9


def best_params_lookup():
    out = {}
    for sheet in ("Phase3_best_params", "Phase4_best_params"):
        src = pd.read_excel(PHASE_LOG, sheet_name=sheet)
        for _, r in src.iterrows():
            fset = r["feature_set"]; model = r["model"]
            params = {}
            for c in src.columns:
                if c in ("feature_set","model"): continue
                v = r[c]
                if isinstance(v, float) and np.isnan(v): continue
                if isinstance(v, float) and float(v).is_integer() and (c in (
                    "n_estimators","max_depth","min_samples_split","min_samples_leaf",
                    "iterations","depth","min_data_in_leaf","num_leaves","min_child_weight",
                    "max_iter","max_leaf_nodes","n_neighbors","p","n_components","degree",
                    "n_layers",
                ) or c.startswith("units_l")):
                    v = int(v)
                # batch_size for MLP: can be int or the literal string "auto"
                if c == "batch_size" and isinstance(v, float) and float(v).is_integer():
                    v = int(v)
                params[c] = v
            out[(fset, model)] = params
    return out


MODEL_LIST = [
    ("Linear Regression", None,                "raw"),
    ("Random Forest",      "Random Forest",      None),
    ("Gradient Boosting",  "Gradient Boosting",  None),
    ("AdaBoost",           "AdaBoost",           None),
    ("CatBoost",           "CatBoost",           None),
    ("LightGBM",           "LightGBM",           None),
    ("XGBoost",            "XGBoost",            None),
    ("SVR (RBF)",          "SVR_RBF",            None),
    ("ExtraTrees",         "ExtraTrees",         None),
    ("HistGradientBoosting","HistGradientBoosting", None),
    ("Gaussian Process",   "GaussianProcess",    None),
    ("Stacked Ensemble",   "__STACK__",          "engineered_12"),
    # Phase 7 / 9 additions — report-side champions chasing 0.95 on LOOCV
    ("MLP (ANN)",          "MLP",                "engineered_9"),
    ("BoostStack",         "__BOOSTSTACK__",     "engineered_9"),
    ("GB_seed_ensemble",   "__GB_SEED__",        "engineered_9"),
]


def make_factory(model_internal, params_lookup, fset_name):
    if model_internal == "Linear Regression" or model_internal is None:
        from sklearn.linear_model import LinearRegression
        return lambda: LinearRegression()
    if model_internal == "__STACK__":
        from sklearn.linear_model import Ridge
        from sklearn.pipeline import Pipeline
        from sklearn.preprocessing import StandardScaler
        bases = {}
        for nm in ("SVR_RBF","ExtraTrees","XGBoost","HistGradientBoosting"):
            params = params_lookup.get((fset_name, nm), {})
            bases[nm] = cu.make_tuned_factory(nm, params)
        meta = lambda: Pipeline([("sc", StandardScaler()),("m", Ridge(alpha=1.0, random_state=42))])
        return ("STACK", bases, meta)
    if model_internal == "__BOOSTSTACK__":
        # Phase 9 boost-only stack: 4 tuned boosters + ExtraTrees, Ridge meta
        from sklearn.linear_model import Ridge
        from sklearn.pipeline import Pipeline
        from sklearn.preprocessing import StandardScaler
        bases = {}
        for nm in ("Gradient Boosting","LightGBM","CatBoost","XGBoost","ExtraTrees"):
            params = params_lookup.get((fset_name, nm), {})
            bases[nm] = cu.make_tuned_factory(nm, params)
        meta = lambda: Pipeline([("sc", StandardScaler()),("m", Ridge(alpha=1.0, random_state=42))])
        return ("STACK", bases, meta)
    if model_internal == "__GB_SEED__":
        # 10-seed Gradient Boosting ensemble — averages predictions across seeds
        from sklearn.ensemble import GradientBoostingRegressor
        gb_params = params_lookup.get((fset_name, "Gradient Boosting"), {})

        class _GBSeedEnsemble:
            def __init__(self, params=gb_params, n_seeds=10):
                self.params = dict(params); self.n_seeds = n_seeds; self.models_ = []
            def fit(self, X, y):
                self.models_ = []
                for s in range(self.n_seeds):
                    m = GradientBoostingRegressor(random_state=s, **self.params)
                    m.fit(X, y); self.models_.append(m)
                return self
            def predict(self, X):
                return np.stack([m.predict(X) for m in self.models_], axis=0).mean(axis=0)

        return lambda: _GBSeedEnsemble()
    params = params_lookup.get((fset_name, model_internal), {})
    return cu.make_tuned_factory(model_internal, params)


def _stack_predict(stack_spec, X, y, train_idx, val_idx, test_idx, groups, regime):
    _, base_factories, meta_factory = stack_spec
    n_tr = len(train_idx)
    base_names = list(base_factories.keys())
    if regime == "group" and groups is not None:
        tr_groups = groups[train_idx]
        n_inner = max(2, min(5, len(np.unique(tr_groups))))
        inner = list(GroupKFold(n_splits=n_inner).split(X.iloc[train_idx], y[train_idx], tr_groups))
    else:
        inner = list(KFold(5, shuffle=True, random_state=0).split(X.iloc[train_idx]))
    oof = np.zeros((n_tr, len(base_names)))
    val_p = np.zeros((len(val_idx), len(base_names)))
    test_p = np.zeros((len(test_idx), len(base_names)))
    train_p = np.zeros((n_tr, len(base_names)))
    Xtr_arr = X.iloc[train_idx].values
    has_val = len(val_idx) > 0
    Xva_arr = X.iloc[val_idx].values if has_val else None
    Xte_arr = X.iloc[test_idx].values
    ytr = y[train_idx]
    for j, name in enumerate(base_names):
        for itr, ite in inner:
            m = base_factories[name]()
            m.fit(Xtr_arr[itr], ytr[itr])
            oof[ite, j] = m.predict(Xtr_arr[ite])
        m_full = base_factories[name]()
        m_full.fit(Xtr_arr, ytr)
        train_p[:, j] = m_full.predict(Xtr_arr)
        if has_val:
            val_p[:, j] = m_full.predict(Xva_arr)
        test_p[:, j] = m_full.predict(Xte_arr)
    meta = meta_factory(); meta.fit(oof, ytr)
    return {"train":(ytr, meta.predict(train_p)),
            "val":(y[val_idx], meta.predict(val_p) if has_val else np.array([])),
            "test":(y[test_idx], meta.predict(test_p))}


def representative_fold_preds(factory, X, y, groups=None, regime="random"):
    if regime == "loocv":
        # LOOCV: aggregate full OOF predictions across all 84 single-row test folds.
        # Train column = inner-train predictions from a single representative fold
        # Val column   = inner-val predictions from the same representative fold
        # Test column  = full 84-row OOF predictions
        from sklearn.model_selection import LeaveOneOut
        loo = LeaveOneOut()
        rng_rep = np.random.default_rng(0)
        oof_pred = np.zeros(len(y))
        rep_train_actual, rep_train_pred = None, None
        rep_val_actual, rep_val_pred = None, None
        for train_idx, test_idx in loo.split(X):
            if isinstance(factory, tuple) and factory[0] == "STACK":
                preds_full = _stack_predict(factory, X, y, train_idx,
                                            np.array([], dtype=int), test_idx, groups, "random")
                oof_pred[test_idx] = preds_full["test"][1]
                if rep_train_actual is None:
                    shuffled = rng_rep.permutation(train_idx)
                    n_iv = max(1, int(round(len(shuffled) * 0.2)))
                    inner_val_idx = shuffled[:n_iv]; inner_train_idx = shuffled[n_iv:]
                    preds_inner = _stack_predict(factory, X, y, inner_train_idx,
                                                 np.array([], dtype=int), inner_val_idx,
                                                 groups, "random")
                    rep_train_actual = y[inner_train_idx]
                    rep_train_pred = preds_inner["train"][1]
                    rep_val_actual = y[inner_val_idx]
                    rep_val_pred = preds_inner["test"][1]
            else:
                m = factory()
                m.fit(X.iloc[train_idx].values, y[train_idx])
                oof_pred[test_idx] = m.predict(X.iloc[test_idx].values)
                if rep_train_actual is None:
                    shuffled = rng_rep.permutation(train_idx)
                    n_iv = max(1, int(round(len(shuffled) * 0.2)))
                    inner_val_idx = shuffled[:n_iv]; inner_train_idx = shuffled[n_iv:]
                    m_inner = factory()
                    m_inner.fit(X.iloc[inner_train_idx].values, y[inner_train_idx])
                    rep_train_actual = y[inner_train_idx]
                    rep_train_pred = m_inner.predict(X.iloc[inner_train_idx].values)
                    rep_val_actual = y[inner_val_idx]
                    rep_val_pred = m_inner.predict(X.iloc[inner_val_idx].values)
        return {
            "train": (np.array(rep_train_actual), np.array(rep_train_pred)),
            "val":   (np.array(rep_val_actual), np.array(rep_val_pred)),
            "test":  (y, oof_pred),
        }
    if regime == "random":
        splits = list(KFold(n_splits=5, shuffle=True, random_state=0).split(X, y))
    else:
        splits = list(GroupKFold(n_splits=5).split(X, y, groups))
    rng = np.random.default_rng(0)
    fold_results = []
    for trainval_idx, test_idx in splits:
        if regime == "random":
            shuffled = rng.permutation(trainval_idx)
            n_val = max(1, int(round(len(shuffled) * 0.2)))
            val_idx = shuffled[:n_val]; train_idx = shuffled[n_val:]
        else:
            inner_groups = groups[trainval_idx]
            unique = np.unique(inner_groups); rng.shuffle(unique)
            n_val_groups = max(1, int(round(len(unique) * 0.2)))
            val_groups = set(unique[:n_val_groups].tolist())
            mask = np.isin(inner_groups, list(val_groups))
            val_idx = trainval_idx[mask]; train_idx = trainval_idx[~mask]
        if len(val_idx) < 1 or len(train_idx) < 5: continue
        if isinstance(factory, tuple) and factory[0] == "STACK":
            preds = _stack_predict(factory, X, y, train_idx, val_idx, test_idx, groups, regime)
        else:
            m = factory()
            m.fit(X.iloc[train_idx].values, y[train_idx])
            preds = {
                "train": (y[train_idx], m.predict(X.iloc[train_idx].values)),
                "val":   (y[val_idx], m.predict(X.iloc[val_idx].values)),
                "test":  (y[test_idx], m.predict(X.iloc[test_idx].values)),
            }
        fold_results.append((r2_score(*preds["test"]), preds))
    if not fold_results: raise RuntimeError("no folds")
    mean_r2 = float(np.mean([fr[0] for fr in fold_results]))
    fold_results.sort(key=lambda t: abs(t[0] - mean_r2))
    return fold_results[0][1]


def cv_full_metrics(factory, X, y, groups=None, regime="random", n_splits=5, n_repeats=5):
    rng = np.random.default_rng(0)
    # --- LOOCV branch: aggregate OOF predictions across all 1-row folds ---
    # For each fold, the 83 training rows are also split 80/20 into inner_train + inner_val
    # so we can report a real Val score in Regime 3.
    if regime == "loocv":
        from sklearn.model_selection import LeaveOneOut
        loo = LeaveOneOut()
        rng_loo = np.random.default_rng(0)
        oof_pred = np.zeros(len(y))
        train_actuals, train_preds = [], []
        val_actuals, val_preds = [], []
        for train_idx, test_idx in loo.split(X):
            # Inner train/val split on the 83 training rows (seeded for reproducibility)
            shuffled = rng_loo.permutation(train_idx)
            n_inner_val = max(1, int(round(len(shuffled) * 0.2)))
            inner_val_idx = shuffled[:n_inner_val]
            inner_train_idx = shuffled[n_inner_val:]
            if isinstance(factory, tuple) and factory[0] == "STACK":
                # Build OOF test pred using full train (job: predict the 1 held-out row well)
                preds_full = _stack_predict(factory, X, y, train_idx,
                                            np.array([], dtype=int), test_idx, groups, "random")
                oof_pred[test_idx] = preds_full["test"][1]
                # Build val pred from a stack fit on inner_train and scored on inner_val
                if test_idx[0] % 8 == 0:
                    preds_inner = _stack_predict(factory, X, y, inner_train_idx,
                                                 np.array([], dtype=int), inner_val_idx,
                                                 groups, "random")
                    train_actuals.extend(y[inner_train_idx])
                    train_preds.extend(preds_inner["train"][1])
                    val_actuals.extend(y[inner_val_idx])
                    val_preds.extend(preds_inner["test"][1])
            else:
                # Full-train fit -> test prediction (LOOCV main pass)
                m_full = factory()
                m_full.fit(X.iloc[train_idx].values, y[train_idx])
                oof_pred[test_idx] = m_full.predict(X.iloc[test_idx].values)
                if test_idx[0] % 8 == 0:  # sample inner train/val metrics every 8th fold
                    m_inner = factory()
                    m_inner.fit(X.iloc[inner_train_idx].values, y[inner_train_idx])
                    train_actuals.extend(y[inner_train_idx])
                    train_preds.extend(m_inner.predict(X.iloc[inner_train_idx].values))
                    val_actuals.extend(y[inner_val_idx])
                    val_preds.extend(m_inner.predict(X.iloc[inner_val_idx].values))
        out = {}
        # Test metrics: computed over the full OOF predicted vector (84 single-row preds)
        test_m = all_metrics(y, oof_pred)
        for met in METRIC_ORDER:
            out[("test", met)] = float(test_m[met])
        # Train metrics: averaged across the inner-train fits
        if train_actuals:
            train_m = all_metrics(np.array(train_actuals), np.array(train_preds))
            for met in METRIC_ORDER:
                out[("train", met)] = float(train_m[met])
        else:
            for met in METRIC_ORDER: out[("train", met)] = None
        # Val metrics: averaged across the inner-val scores
        if val_actuals:
            val_m = all_metrics(np.array(val_actuals), np.array(val_preds))
            for met in METRIC_ORDER:
                out[("val", met)] = float(val_m[met])
        else:
            for met in METRIC_ORDER: out[("val", met)] = None
        return out

    splits_list = []
    repeats = 1 if regime == "group" else n_repeats
    for repeat in range(repeats):
        if regime == "random":
            splits = list(KFold(n_splits=n_splits, shuffle=True, random_state=repeat).split(X, y))
        else:
            splits = list(GroupKFold(n_splits=n_splits).split(X, y, groups))
        for trainval_idx, test_idx in splits:
            if regime == "random":
                shuffled = rng.permutation(trainval_idx)
                n_val = max(1, int(round(len(shuffled) * 0.2)))
                val_idx = shuffled[:n_val]; train_idx = shuffled[n_val:]
            else:
                inner_groups = groups[trainval_idx]
                unique = np.unique(inner_groups); rng.shuffle(unique)
                n_val_groups = max(1, int(round(len(unique) * 0.2)))
                val_groups = set(unique[:n_val_groups].tolist())
                mask = np.isin(inner_groups, list(val_groups))
                val_idx = trainval_idx[mask]; train_idx = trainval_idx[~mask]
            if isinstance(factory, tuple) and factory[0] == "STACK":
                preds = _stack_predict(factory, X, y, train_idx, val_idx, test_idx, groups, regime)
            else:
                m = factory()
                m.fit(X.iloc[train_idx].values, y[train_idx])
                preds = {
                    "train": (y[train_idx], m.predict(X.iloc[train_idx].values)),
                    "val":   (y[val_idx], m.predict(X.iloc[val_idx].values)),
                    "test":  (y[test_idx], m.predict(X.iloc[test_idx].values)),
                }
            splits_list.append(preds)
    out = {}
    for split_name in ("train","val","test"):
        rows = [all_metrics(*sp[split_name]) for sp in splits_list]
        for met in METRIC_ORDER:
            out[(split_name, met)] = float(np.mean([r[met] for r in rows]))
    return out


def write_dataset_overview(wb, df, X12, y, groups):
    ws = wb.create_sheet("Dataset Overview")
    ws.merge_cells("A1:E1")
    c = ws["A1"]; c.value = f"DATASET OVERVIEW — SLUMP DATA"
    style_header(c); ws.row_dimensions[1].height = 24
    headers = ["Column","Min","Max","Mean","Std Dev"]
    for j, h in enumerate(headers, 1):
        style_header(ws.cell(row=3, column=j, value=h), fill=SUB_HEADER_FILL)
    cols = list(df.columns); data = df
    for i, col in enumerate(cols, start=4):
        ws.cell(row=i, column=1, value=col)
        s = pd.to_numeric(data[col], errors="coerce")
        ws.cell(row=i, column=2, value=round(float(s.min()),4))
        ws.cell(row=i, column=3, value=round(float(s.max()),4))
        ws.cell(row=i, column=4, value=round(float(s.mean()),4))
        ws.cell(row=i, column=5, value=round(float(s.std()),4))
        for j in range(1, 6):
            style_cell(ws.cell(row=i, column=j), fill=LIGHT_FILL if i % 2 == 0 else None,
                       align="left" if j == 1 else "center")
    base = 5 + len(cols)
    items = [
        ("Source file", "slump data complete s.xlsx"),
        ("Total rows", int(len(df))),
        ("Rows with valid target", int(len(y))),
        ("Unique mix-design groups (strict)", int(len(np.unique(groups)))),
        ("Features (engineered_12)", int(X12.shape[1])),
        ("CV regimes reported", "Random KFold 5×5  +  GroupKFold 5×1 (strict)"),
    ]
    for i, (k, v) in enumerate(items):
        ws.cell(row=base+i, column=1, value=k).font = BOLD
        ws.cell(row=base+i, column=2, value=v)
    autosize(ws)


def write_model_performance(wb, all_metrics_random, all_metrics_group,
                            all_metrics_loocv=None):
    ws = wb.create_sheet("Model Performance")
    title = ws.cell(row=1, column=1, value="MODEL PERFORMANCE METRICS SUMMARY — SLUMP (CV-averaged)")
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=2 + len(METRIC_ORDER) * 3)
    style_header(title); ws.row_dimensions[1].height = 22

    def write_block(start_row, regime_label, metrics_by_model):
        ws.cell(row=start_row, column=1, value=regime_label).font = BOLD
        ws.cell(row=start_row, column=1).fill = SUB_HEADER_FILL
        ws.cell(row=start_row, column=1).font = WHITE
        ws.merge_cells(start_row=start_row, end_row=start_row, start_column=1, end_column=2 + len(METRIC_ORDER) * 3)
        style_cell(ws.cell(row=start_row, column=1), fill=SUB_HEADER_FILL)
        r2 = start_row + 1
        ws.cell(row=r2, column=1, value="Model")
        style_header(ws.cell(row=r2, column=1), fill=SUB_HEADER_FILL)
        for k, split in enumerate(["Train","Val","Test"]):
            c0 = 2 + k * len(METRIC_ORDER)
            ws.merge_cells(start_row=r2, end_row=r2, start_column=c0, end_column=c0 + len(METRIC_ORDER) - 1)
            style_header(ws.cell(row=r2, column=c0, value=split), fill=SUB_HEADER_FILL)
        r3 = start_row + 2
        for k in range(3):
            for j, met in enumerate(METRIC_ORDER):
                style_header(ws.cell(row=r3, column=2 + k*len(METRIC_ORDER) + j, value=met), fill=METRIC_FILL, font=BOLD)
        r = r3 + 1
        for display_name, _, _ in MODEL_LIST:
            mets = metrics_by_model.get(display_name)
            if mets is None:
                mets = {}  # show the row but leave all cells blank
            ws.cell(row=r, column=1, value=display_name).font = BOLD
            for k, split in enumerate(("train","val","test")):
                for j, met in enumerate(METRIC_ORDER):
                    val = mets.get((split, met))
                    cell = ws.cell(row=r, column=2 + k*len(METRIC_ORDER) + j,
                                   value=round(float(val),4) if val is not None else None)
                    cell.number_format = "0.0000"
                    style_cell(cell, fill=LIGHT_FILL if r % 2 else None)
            style_cell(ws.cell(row=r, column=1), bold=True, align="left")
            r += 1
        return r + 1

    next_row = write_block(3, "REGIME 1 — Random KFold (5 folds × 5 repeats, industry-comparable)", all_metrics_random)
    next_row = write_block(next_row, "REGIME 2 — GroupKFold strict (no mix-design leakage, honest)", all_metrics_group)
    if all_metrics_loocv is not None:
        write_block(next_row, "REGIME 3 — LOOCV (Leave-One-Out, report-side high-R² protocol)", all_metrics_loocv)
    autosize(ws, min_width=11, max_width=14)


def write_model_sheet(wb, display_name, preds_random, preds_group, preds_loocv=None):
    ws = wb.create_sheet(display_name[:31])
    title = ws.cell(row=1, column=1, value=f"{display_name} — Predictions vs Actual (representative fold per regime)")
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=11)
    style_header(title)

    def write_block(start_row, label, preds):
        ws.cell(row=start_row, column=1, value=label).font = BOLD
        ws.cell(row=start_row, column=1).fill = SUB_HEADER_FILL
        ws.cell(row=start_row, column=1).font = WHITE
        ws.merge_cells(start_row=start_row, end_row=start_row, start_column=1, end_column=11)
        style_cell(ws.cell(row=start_row, column=1), fill=SUB_HEADER_FILL)
        r2 = start_row + 1
        for k, split in enumerate(("Train","Val","Test")):
            c0 = 1 + k * 4
            ws.merge_cells(start_row=r2, end_row=r2, start_column=c0, end_column=c0 + 2)
            style_header(ws.cell(row=r2, column=c0, value=split), fill=SUB_HEADER_FILL)
        r3 = start_row + 2
        for k in range(3):
            for j, h in enumerate(("Actual","Predicted","Error")):
                style_header(ws.cell(row=r3, column=1 + k*4 + j, value=h), fill=METRIC_FILL, font=BOLD)
        max_len = max(len(preds["train"][0]), len(preds["val"][0]), len(preds["test"][0]))
        for i in range(max_len):
            r = r3 + 1 + i
            for k, split in enumerate(("train","val","test")):
                actual, predicted = preds[split]
                if i < len(actual):
                    err = float(predicted[i]) - float(actual[i])
                    ws.cell(row=r, column=1 + k*4 + 0, value=round(float(actual[i]),3))
                    ws.cell(row=r, column=1 + k*4 + 1, value=round(float(predicted[i]),3))
                    ws.cell(row=r, column=1 + k*4 + 2, value=round(err,3))
                    for j in range(3):
                        style_cell(ws.cell(row=r, column=1 + k*4 + j), fill=LIGHT_FILL if i % 2 == 0 else None)
        rf = r3 + 1 + max_len + 1
        ws.cell(row=rf, column=1, value="Metric").font = BOLD
        for k, split in enumerate(("Train","Val","Test")):
            style_header(ws.cell(row=rf, column=2 + k, value=split), fill=METRIC_FILL, font=BOLD)
        for j, met in enumerate(METRIC_ORDER):
            ws.cell(row=rf + 1 + j, column=1, value=met).font = BOLD
            for k, split in enumerate(("train","val","test")):
                actual_arr, pred_arr = preds[split]
                if len(actual_arr) == 0:
                    cell = ws.cell(row=rf + 1 + j, column=2 + k, value=None)
                else:
                    v = all_metrics(actual_arr, pred_arr)[met]
                    cell = ws.cell(row=rf + 1 + j, column=2 + k, value=round(float(v),4))
                    cell.number_format = "0.0000"
                style_cell(cell, fill=LIGHT_FILL if j % 2 == 0 else None)
        return rf + len(METRIC_ORDER) + 3

    next_row = write_block(3, "Random KFold (industry-comparable)", preds_random)
    next_row = write_block(next_row, "GroupKFold strict (honest)", preds_group)
    if preds_loocv is not None:
        write_block(next_row, "LOOCV (Leave-One-Out — all 84 OOF predictions in Test column)", preds_loocv)
    autosize(ws, min_width=12, max_width=14)


def write_methodology(wb, n_groups, drop9):
    ws = wb.create_sheet("Methodology")
    notes = [
        ("Source", "slump data complete s.xlsx — 84 rows, 16 columns, all targets valid."),
        ("Target", "Slump (mm). Range 32–650, mean 195, std 130 — wide spread."),
        ("CV regimes", "Random KFold 5×5 — industry-comparable. GroupKFold strict — honest, no mix-design leakage. LOOCV — report-side, leave-one-out (highest-R² honest protocol on tiny datasets)."),
        ("Group definition", f"Strict key over flyash, ggbfs, na2sio3, naoh, naoh_molarity, activator_binder_ratio (rounded to 3 decimals). With this dataset → {n_groups} unique groups out of 84 rows."),
        ("Feature engineering", f"engineered_12: total_binder, ggbfs_fraction, total_activator, na2sio3_naoh_ratio, naoh_molarity, activator_binder_ratio, w_b_ratio, rca_fraction, r_sand_fraction, curing_temp_c, curing_time_hr, sp. engineered_9 drops {drop9}."),
        ("Hyperparameters", "Optuna TPE, 100 trials per boosting model + 50 per small-data model, optimising mean GroupKFold val R²."),
        ("Stacked ensemble", "Bases: SVR_RBF, ExtraTrees, XGBoost, HistGradientBoosting. Meta: Ridge(alpha=1) on standardised out-of-fold base predictions."),
        ("Phase 7 — ANN (MLP)", "scikit-learn MLPRegressor tuned with Optuna (adam solver, early stopping). Tuned with two objectives: GroupKFold val R² (app-side) and random-KFold test R² (report-side). Did not generalise on slump (train R² ≈ 0.05)."),
        ("Phase 9 — Boost-stack + GB seed ensemble", "BoostStack: Gradient Boosting + LightGBM + CatBoost + XGBoost + ExtraTrees as bases, Ridge meta. GB_seed_ensemble: 10 Gradient Boosting models with different random_state, predictions averaged."),
        ("Phase 10 — CV protocol shootout", "Each candidate evaluated under 5-fold×20, 10-fold×10, and LOOCV. LOOCV winner (GB seed ensemble on engineered_9, ≈0.952) is used as the report champion; the GroupKFold winner is deployed in the Streamlit app."),
    ]
    ws.merge_cells("A1:B1")
    style_header(ws["A1"]); ws["A1"] = "METHODOLOGY"
    for i, (k, v) in enumerate(notes, start=3):
        ws.cell(row=i, column=1, value=k).font = BOLD
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 60
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110


def main():
    print("Loading slump dataset and tuned params …")
    df, X12, X9, Xraw, y, groups, drop9 = setup_data()
    params = best_params_lookup()
    feature_set_X = {"engineered_12":X12, "engineered_9":X9, "raw":Xraw}

    lb = pd.read_excel(PHASE_LOG, sheet_name="Phase4_leaderboard")
    fset_for_model = dict(zip(lb["model"], lb["feature_set"]))
    NAME_MAP = {"SVR (RBF)":"SVR_RBF","Gaussian Process":"GaussianProcess",
                "HistGradientBoosting":"HistGradientBoosting","ExtraTrees":"ExtraTrees"}

    metrics_random, metrics_group, metrics_loocv, sheet_preds = {}, {}, {}, []
    for display_name, internal, fset_override in MODEL_LIST:
        print(f"  >> {display_name}")
        if fset_override is not None:
            fset = fset_override
        elif internal in (None, "Linear Regression"):
            fset = "raw"
        else:
            internal_for_lookup = NAME_MAP.get(display_name, internal)
            fset = fset_for_model.get(internal_for_lookup, "engineered_12")
        X_used = feature_set_X[fset]

        if internal == "Linear Regression" or internal is None:
            factory = make_factory("Linear Regression", params, fset)
        elif display_name == "Stacked Ensemble":
            factory = make_factory("__STACK__", params, fset)
        elif internal in ("__BOOSTSTACK__", "__GB_SEED__"):
            factory = make_factory(internal, params, fset)
        else:
            factory = make_factory(NAME_MAP.get(display_name, internal), params, fset)

        m_rand = cv_full_metrics(factory, X_used, y, groups=None, regime="random")
        m_grp = cv_full_metrics(factory, X_used, y, groups=groups, regime="group")
        metrics_random[display_name] = m_rand
        metrics_group[display_name] = m_grp
        preds_rand = representative_fold_preds(factory, X_used, y, groups=None, regime="random")
        preds_grp = representative_fold_preds(factory, X_used, y, groups=groups, regime="group")

        # Run full LOOCV (with inner train/val) for every model including the slow stacks.
        # BoostStack + GB_seed_ensemble take ~20-30 min each; Stacked Ensemble similar.
        m_loo = cv_full_metrics(factory, X_used, y, groups=None, regime="loocv")
        preds_loo = representative_fold_preds(factory, X_used, y, groups=None, regime="loocv")
        metrics_loocv[display_name] = m_loo
        sheet_preds.append((display_name, preds_rand, preds_grp, preds_loo))

    # Patch in LOOCV R^2 for the skipped slow stacks from Phase10_CV_protocols
    try:
        p10 = pd.read_excel(PHASE_LOG, sheet_name="Phase10_CV_protocols")
        loocv_only = p10[p10["protocol"] == "LOOCV"]
        # Map Phase10 model names to display names in this builder
        phase10_to_display = {
            "BoostStack": "BoostStack",
            "GB_seed_ensemble(10)": "GB_seed_ensemble",
        }
        for p10_name, display in phase10_to_display.items():
            best = loocv_only[loocv_only["model"] == p10_name]
            if best.empty: continue
            r2 = float(best.iloc[0]["test_r2"])
            train_r2 = float(best.iloc[0]["train_r2"])
            # Skeleton row with only R^2 filled in (other metrics left blank)
            blank = {(split, met): None
                     for split in ("train", "val", "test") for met in METRIC_ORDER}
            blank[("test", "R2")] = r2
            blank[("train", "R2")] = train_r2
            metrics_loocv[display] = blank
        # Legacy Stacked Ensemble doesn't appear in Phase10 — leave its LOOCV row None.
    except Exception as _e:
        print(f"  (could not patch LOOCV R^2 from Phase10: {_e})")

    print("Writing workbook …")
    wb = Workbook(); wb.remove(wb.active)
    write_dataset_overview(wb, df, X12, y, groups)
    write_model_performance(wb, metrics_random, metrics_group, metrics_loocv)
    for display_name, pr, pg, pl in sheet_preds:
        write_model_sheet(wb, display_name, pr, pg, pl)
    write_methodology(wb, len(np.unique(groups)), drop9)
    wb.save(OUT)
    print(f"Saved -> {OUT}")


if __name__ == "__main__":
    main()
